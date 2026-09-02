"""Generate high-cardinality administrative geography source packages.

The county packages are generated from the rows in their registered artifacts.
This keeps the checked-in fixture packages small while letting the exact same
workflow expand to every publisher county row after integration fetches the
real bytes. The generator only maps source columns to Ledger source selectors;
it does not reconcile, age, impute, or otherwise derive fact values.
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
import re
from typing import Any, Callable
from zipfile import ZipFile

from openpyxl import load_workbook
import yaml

from chronicle.epoch import schema_id
from chronicle.sources.cells import decode_delimited_text


REPO_ROOT = Path(__file__).resolve().parents[2]
GENERATED_BY = "uv run python -m chronicle.sources.admin_packages"

IRS_PACKAGE_ID = "soi-county-2022"
PEP_PACKAGE_ID = "census-pep-county-population-2024"
SNAP_PACKAGE_ID = "usda-snap-fy2025-monthly-state-caseloads"

IRS_PACKAGE_PATH = Path("packages/irs_soi/county_2022/source_package.yaml")
PEP_PACKAGE_PATH = Path("packages/census/pep_county_2024/source_package.yaml")
SNAP_PACKAGE_PATH = Path(
    "packages/usda_snap/fy2025_monthly_state_caseloads/source_package.yaml"
)


class _NoAliasDumper(yaml.SafeDumper):
    """Keep generated package YAML explicit instead of emitting anchors."""

    def ignore_aliases(self, data: Any) -> bool:
        return True


def build_irs_soi_county_package(root: Path = REPO_ROOT) -> dict[str, Any]:
    """Build the TY2022 IRS SOI county package mapping from its CSV rows."""
    data_dir = root / "db/data/irs_soi/county_2022"
    manifest = _load_yaml(data_dir / "manifest.yaml")
    artifact = _year_file(manifest, 2022)
    fixture = bool(artifact.get("fixture"))
    headers, source_rows = _read_csv(data_dir / artifact["filename"])
    _require_headers(
        headers,
        "STATEFIPS",
        "STATE",
        "COUNTYFIPS",
        "COUNTYNAME",
        "N1",
        "A00100",
    )

    selected_rows: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []
    seen_fips: set[str] = set()
    eligible = []
    for source_row in source_rows:
        county_fips = _county_fips(source_row["STATEFIPS"], source_row["COUNTYFIPS"])
        if county_fips.endswith("000"):
            continue
        if county_fips.endswith("999"):
            raise ValueError(f"Reserved IRS aggregate FIPS row: {county_fips}")
        if "agi_stub" in source_row and _selector_scalar(source_row["agi_stub"]) != 0:
            continue
        _require_numeric(source_row, "N1", "A00100")
        if county_fips in seen_fips:
            raise ValueError(f"Duplicate IRS county FIPS row: {county_fips}")
        seen_fips.add(county_fips)
        eligible.append((county_fips, source_row))
    if not eligible:
        raise ValueError("IRS SOI source has no eligible county rows")
    if not fixture and not 3_000 <= len(eligible) <= 4_000:
        raise ValueError(
            f"Implausible IRS SOI county inventory: {len(eligible)} eligible rows"
        )

    state_column = _excel_column(headers.index("STATEFIPS") + 1)
    county_column = _excel_column(headers.index("COUNTYFIPS") + 1)
    name_column = _excel_column(headers.index("COUNTYNAME") + 1)
    for ordinal, (county_fips, source_row) in enumerate(sorted(eligible)):
        criteria = {
            "STATEFIPS": _selector_scalar(source_row["STATEFIPS"]),
            "COUNTYFIPS": _selector_scalar(source_row["COUNTYFIPS"]),
        }
        if "agi_stub" in source_row:
            criteria["agi_stub"] = _selector_scalar(source_row["agi_stub"])
        selected_rows.append(criteria)
        county_name = source_row["COUNTYNAME"].strip()
        rows.append(
            {
                "value_id": f"county_{county_fips}",
                "label": county_name,
                "ordinal": ordinal,
                "row_number": ordinal + 2,
                "geography_id": f"0500000US{county_fips}",
                "geography_level": "county",
                "geography_name": county_name,
                "geography_vintage": "2022",
                "expected_row_header_column": name_column,
                "expected_row_header": county_name,
                "filters": {"filing_status": "all", "income_range": "all"},
                "guard_cells": [
                    {
                        "column": state_column,
                        "row": "start",
                        "expected_value": criteria["STATEFIPS"],
                        "label": "state FIPS",
                    },
                    {
                        "column": county_column,
                        "row": "start",
                        "expected_value": criteria["COUNTYFIPS"],
                        "label": "county FIPS",
                    },
                ],
                "table_record_kind": "total",
            }
        )

    fixture_prefix = "TEST FIXTURE — " if fixture else ""
    source_name = "test_fixture_irs_soi" if fixture else "irs_soi"
    extraction_method = (
        "TEST FIXTURE ONLY — synthetic county rows, not IRS data; full CSV "
        "row parse with selected county source-row lineage"
        if fixture
        else "full publisher CSV row parse with selected county source-row lineage"
    )
    record_set_id = "irs_soi.ty2022.county_totals"
    return {
        "schema_version": schema_id("source_package"),
        "package_id": IRS_PACKAGE_ID,
        "label": f"{fixture_prefix}IRS SOI 2022 county return and AGI totals",
        "fixture": fixture,
        "generated_by": GENERATED_BY,
        "artifact": {
            "source_name": source_name,
            "source_table": f"{fixture_prefix}IRS SOI County Data 2022",
            "resource_package": "db",
            "resource_directory": "data/irs_soi/county_2022",
            "manifest": "manifest.yaml",
            "vintage": "tax_year_2022",
            "extracted_at": _artifact_extracted_date(artifact),
            "extraction_method": extraction_method,
            "parser": "delimited_text_full_rows",
            "sheet_name": "22incyallnoagi",
            "selected_rows": selected_rows,
            **({} if fixture else {"artifact_year": 2022}),
        },
        "record_sets": [
            {
                "record_set_id": record_set_id,
                "provenance_class": "administrative",
                "record_set_spec_id": "irs_soi.county_totals.v1",
                "source_record_id_prefix": record_set_id,
                "sheet_name": "22incyallnoagi",
                "period_type": "tax_year",
                "period": "2022",
                "geography_id": "0100000US",
                "geography_level": "country",
                "geography_name": "United States",
                "geography_vintage": "2022",
                "entity": "tax_unit",
                "entity_role": "filing_unit",
                "domain": "all_individual_income_tax_returns",
                "groupby_dimension": "irs_soi.county",
                "rows": rows,
                "measures": [
                    {
                        "measure_id": "return_count",
                        "label": "Number of returns",
                        "ordinal": 0,
                        "column": _excel_column(headers.index("N1") + 1),
                        "source_column_id": "N1",
                        "expected_column_header_row": 1,
                        "expected_column_header": "N1",
                        "concept": "irs_soi.individual_income_tax_returns",
                        "unit": "count",
                        "aggregation": "sum",
                        "expected_cell_type": "number",
                    },
                    {
                        "measure_id": "adjusted_gross_income",
                        "label": "Adjusted gross income",
                        "ordinal": 1,
                        "column": _excel_column(headers.index("A00100") + 1),
                        "source_column_id": "A00100",
                        "expected_column_header_row": 1,
                        "expected_column_header": "A00100",
                        "concept": "us:statutes/26/62#adjusted_gross_income",
                        "source_concept": "irs_soi.adjusted_gross_income",
                        "concept_relation": "exact",
                        "concept_authority": "ledger-us",
                        "concept_evidence_url": (
                            "https://uscode.house.gov/view.xhtml?"
                            "req=(title:26%20section:62%20edition:prelim)"
                        ),
                        "concept_evidence_notes": (
                            "IRS SOI county files report adjusted gross income "
                            "for individual income tax returns; IRC section 62 "
                            "defines adjusted gross income."
                        ),
                        "legal_vintage": "tax_year_2022",
                        "unit": "usd",
                        "aggregation": "sum",
                        "value_scale": 1000,
                        "expected_cell_type": "number",
                    },
                ],
            }
        ],
    }


def build_census_pep_county_package(root: Path = REPO_ROOT) -> dict[str, Any]:
    """Build the Vintage 2024 PEP county package mapping from its CSV rows."""
    data_dir = root / "db/data/census/pep_county_2024"
    manifest = _load_yaml(data_dir / "manifest.yaml")
    artifact = _year_file(manifest, 2024)
    fixture = bool(artifact.get("fixture"))
    headers, source_rows = _read_csv(data_dir / artifact["filename"])
    _require_headers(
        headers,
        "SUMLEV",
        "STATE",
        "COUNTY",
        "STNAME",
        "CTYNAME",
        "POPESTIMATE2024",
    )

    selected_rows: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []
    seen_fips: set[str] = set()
    eligible = []
    for source_row in source_rows:
        if _selector_scalar(source_row["SUMLEV"]) != 50:
            continue
        county_fips = _county_fips(source_row["STATE"], source_row["COUNTY"])
        if county_fips.endswith("000"):
            continue
        _require_numeric(source_row, "POPESTIMATE2024")
        if county_fips in seen_fips:
            raise ValueError(f"Duplicate Census PEP county FIPS row: {county_fips}")
        seen_fips.add(county_fips)
        eligible.append((county_fips, source_row))
    if not eligible:
        raise ValueError("Census PEP source has no eligible county rows")
    if not fixture and not 3_000 <= len(eligible) <= 4_000:
        raise ValueError(
            f"Implausible Census PEP county inventory: {len(eligible)} eligible rows"
        )

    sumlev_column = _excel_column(headers.index("SUMLEV") + 1)
    state_column = _excel_column(headers.index("STATE") + 1)
    county_column = _excel_column(headers.index("COUNTY") + 1)
    name_column = _excel_column(headers.index("CTYNAME") + 1)
    for ordinal, (county_fips, source_row) in enumerate(sorted(eligible)):
        criteria = {
            "SUMLEV": _selector_scalar(source_row["SUMLEV"]),
            "STATE": _selector_scalar(source_row["STATE"]),
            "COUNTY": _selector_scalar(source_row["COUNTY"]),
        }
        selected_rows.append(criteria)
        county_name = source_row["CTYNAME"].strip()
        rows.append(
            {
                "value_id": f"county_{county_fips}",
                "label": county_name,
                "ordinal": ordinal,
                "row_number": ordinal + 2,
                "geography_id": f"0500000US{county_fips}",
                "geography_level": "county",
                "geography_name": county_name,
                "geography_vintage": "2024",
                "expected_row_header_column": name_column,
                "expected_row_header": county_name,
                "guard_cells": [
                    {
                        "column": sumlev_column,
                        "row": "start",
                        "expected_value": criteria["SUMLEV"],
                        "label": "county summary level",
                    },
                    {
                        "column": state_column,
                        "row": "start",
                        "expected_value": criteria["STATE"],
                        "label": "state FIPS",
                    },
                    {
                        "column": county_column,
                        "row": "start",
                        "expected_value": criteria["COUNTY"],
                        "label": "county FIPS",
                    },
                ],
                "table_record_kind": "total",
            }
        )

    fixture_prefix = "TEST FIXTURE — " if fixture else ""
    source_name = "test_fixture_census_pep" if fixture else "census_pep"
    extraction_method = (
        "TEST FIXTURE ONLY — synthetic county rows, not Census data; full CSV "
        "row parse with selected county source-row lineage"
        if fixture
        else "full publisher CSV row parse with selected county source-row lineage"
    )
    record_set_id = "census_pep.vintage2024.county_population"
    return {
        "schema_version": schema_id("source_package"),
        "package_id": PEP_PACKAGE_ID,
        "label": f"{fixture_prefix}Census PEP Vintage 2024 county population",
        "fixture": fixture,
        "generated_by": GENERATED_BY,
        "artifact": {
            "source_name": source_name,
            "source_table": f"{fixture_prefix}Vintage 2024 County Population Totals",
            "resource_package": "db",
            "resource_directory": "data/census/pep_county_2024",
            "manifest": "manifest.yaml",
            "vintage": "vintage_2024",
            "extracted_at": _artifact_extracted_date(artifact),
            "extraction_method": extraction_method,
            "parser": "delimited_text_full_rows",
            "sheet_name": "co-est2024-alldata",
            "selected_rows": selected_rows,
            **({} if fixture else {"artifact_year": 2024}),
        },
        "record_sets": [
            {
                "record_set_id": record_set_id,
                "provenance_class": "census",
                "record_set_spec_id": "census_pep.county_population.v1",
                "source_record_id_prefix": record_set_id,
                "sheet_name": "co-est2024-alldata",
                "period_type": "calendar_year",
                "period": 2024,
                "geography_id": "0100000US",
                "geography_level": "country",
                "geography_name": "United States",
                "geography_vintage": "2024",
                "entity": "person",
                "entity_role": "resident_population",
                "domain": "resident_population",
                "groupby_dimension": "census_pep.county",
                "rows": rows,
                "measures": [
                    {
                        "measure_id": "resident_population",
                        "label": "Resident population",
                        "ordinal": 0,
                        "column": _excel_column(headers.index("POPESTIMATE2024") + 1),
                        "source_column_id": "POPESTIMATE2024",
                        "expected_column_header_row": 1,
                        "expected_column_header": "POPESTIMATE2024",
                        "concept": "census_pep.resident_population",
                        "source_concept": "census_pep.POPESTIMATE2024",
                        "concept_relation": "source_label",
                        "concept_authority": "census",
                        "unit": "count",
                        "aggregation": "sum",
                        "expected_cell_type": "number",
                    }
                ],
            }
        ],
    }


def build_usda_snap_monthly_package(root: Path = REPO_ROOT) -> dict[str, Any]:
    """Build FY2025 monthly SNAP state caseload selectors from FNS cells."""
    data_dir = root / "db/data/usda_snap/fy69_to_current"
    manifest_name = "manifest_fy2025_monthly_source_package.yaml"
    manifest = _load_yaml(data_dir / manifest_name)
    artifact = _year_file(manifest, 2025)
    archive_path = data_dir / artifact["filename"]
    with ZipFile(archive_path) as archive:
        workbook_content = archive.read("FY25.xlsx")
    workbook = load_workbook(BytesIO(workbook_content), read_only=True, data_only=True)

    existing = _load_yaml(
        root / "packages/usda_snap/fy69_to_current/source_package.yaml"
    )
    regional_sets = {
        item["sheet_name"].lower(): item
        for item in existing["record_sets"]
        if ".state_average_monthly_households." in item["record_set_id"]
    }
    sheet_order = ["nero", "maro", "sero", "mwro", "swro", "mpro", "wro"]
    if set(regional_sets) != set(sheet_order):
        raise ValueError("Unexpected FY2024 SNAP regional source-package inventory")

    expected_title = (
        "SNAP Monthly State Participation and Benefit Summary "
        "P-EBT/Other Excluded - Public Data, Fiscal Year 2025"
    )
    entity_specs = (
        {
            "suffix": "state_households",
            "entity": "household",
            "entity_role": "snap_household",
            "column": "B",
            "header": "Household",
            "measure_id": "participating_households",
            "measure_label": "Participating households",
            "concept": "usda_snap.monthly_participating_households",
        },
        {
            "suffix": "state_persons",
            "entity": "person",
            "entity_role": "snap_participant",
            "column": "C",
            "header": "Persons",
            "measure_id": "participating_persons",
            "measure_label": "Participating persons",
            "concept": "usda_snap.monthly_participating_persons",
        },
    )

    record_sets = []
    for month_offset in range(1, 13):
        for region in sheet_order:
            template = regional_sets[region]
            worksheet = workbook[template["sheet_name"]]
            if worksheet["A2"].value != expected_title:
                raise ValueError(f"Unexpected FY2025 SNAP title on {worksheet.title}")
            for entity_spec in entity_specs:
                rows = []
                period: str | None = None
                source_month_label: str | None = None
                for template_row in template["rows"]:
                    heading_row = int(template_row["row_number"]) - 13
                    month_row = heading_row + month_offset
                    month_label = worksheet.cell(row=month_row, column=1).value
                    row_period = _snap_month_period(month_label)
                    value_column = 2 if entity_spec["column"] == "B" else 3
                    value = worksheet.cell(row=month_row, column=value_column).value
                    if not isinstance(value, int | float) or isinstance(value, bool):
                        continue
                    if period is not None and period != row_period:
                        raise ValueError(
                            f"Inconsistent SNAP month labels on {worksheet.title}"
                        )
                    period = row_period
                    source_month_label = str(month_label)
                    geography_name = template_row["geography_name"]
                    if (
                        worksheet.cell(row=heading_row, column=1).value
                        != geography_name
                    ):
                        raise ValueError(
                            f"Unexpected geography heading for {geography_name}"
                        )
                    rows.append(
                        {
                            "value_id": template_row["value_id"],
                            "label": geography_name,
                            "ordinal": len(rows),
                            "row_number": month_row,
                            "geography_id": template_row["geography_id"],
                            "geography_level": "state",
                            "geography_name": geography_name,
                            "geography_vintage": "current",
                            "expected_row_header_column": "A",
                            "expected_row_header": source_month_label,
                            "guard_cells": [
                                {
                                    "column": "A",
                                    "row": 2,
                                    "expected_value": expected_title,
                                    "label": "table title",
                                },
                                {
                                    "column": "A",
                                    "row": heading_row,
                                    "expected_value": geography_name,
                                    "label": "geography block",
                                },
                            ],
                            "table_record_kind": "total",
                        }
                    )
                if not rows:
                    continue
                if period is None:
                    raise AssertionError("Nonempty SNAP row set has no period")
                period_id = period.replace("-", "_")
                record_set_id = (
                    f"usda_snap.month{period_id}.{entity_spec['suffix']}.{region}"
                )
                concept = entity_spec["concept"]
                record_sets.append(
                    {
                        "record_set_id": record_set_id,
                        "provenance_class": "administrative",
                        "record_set_spec_id": (
                            f"usda_snap.monthly_{entity_spec['suffix']}.v1"
                        ),
                        "source_record_id_prefix": record_set_id,
                        "sheet_name": template["sheet_name"],
                        "period_type": "month",
                        "period": period,
                        "geography_id": "0100000US",
                        "geography_level": "country",
                        "geography_name": "United States",
                        "geography_vintage": "current",
                        "entity": entity_spec["entity"],
                        "entity_role": entity_spec["entity_role"],
                        "domain": "supplemental_nutrition_assistance_program",
                        "groupby_dimension": "usda_snap.state_or_territory",
                        "rows": rows,
                        "measures": [
                            {
                                "measure_id": entity_spec["measure_id"],
                                "label": entity_spec["measure_label"],
                                "ordinal": 0,
                                "column": entity_spec["column"],
                                "source_column_id": entity_spec["header"],
                                "expected_column_header_row": 7,
                                "expected_column_header": entity_spec["header"],
                                "concept": concept,
                                "source_concept": concept,
                                "concept_relation": "source_label",
                                "concept_authority": "ledger-us",
                                "unit": "count",
                                "aggregation": "sum",
                                "expected_cell_type": "number",
                            }
                        ],
                    }
                )

    if not record_sets:
        raise ValueError("FY2025 SNAP workbook has no numeric monthly caseload rows")
    latest_period = max(record_set["period"] for record_set in record_sets)
    latest_month = datetime.strptime(latest_period, "%Y-%m").strftime("%B %Y")

    return {
        "schema_version": schema_id("source_package"),
        "package_id": SNAP_PACKAGE_ID,
        "label": f"USDA FNS SNAP FY2025 monthly state caseloads through {latest_month}",
        "fixture": False,
        "generated_by": GENERATED_BY,
        "artifact": {
            "source_name": "usda_snap",
            "source_table": "SNAP FY2025 Monthly State Participation",
            "resource_package": "db",
            "resource_directory": "data/usda_snap/fy69_to_current",
            "manifest": manifest_name,
            "vintage": f"fiscal_year_2025_through_{latest_period.replace('-', '_')}",
            "extracted_at": _artifact_extracted_date(artifact),
            "extraction_method": (
                "ZIP archive FY25 XLSX member whole-workbook used-range cell parse"
            ),
            "parser": "zip_xlsx_used_range",
            "archive_member": "FY25.xlsx",
            "artifact_year": 2025,
        },
        "record_sets": record_sets,
    }


PACKAGE_BUILDERS: dict[str, tuple[Path, Callable[[Path], dict[str, Any]]]] = {
    IRS_PACKAGE_ID: (IRS_PACKAGE_PATH, build_irs_soi_county_package),
    PEP_PACKAGE_ID: (PEP_PACKAGE_PATH, build_census_pep_county_package),
    SNAP_PACKAGE_ID: (SNAP_PACKAGE_PATH, build_usda_snap_monthly_package),
}


def render_package(payload: dict[str, Any]) -> str:
    """Render a generated package deterministically."""
    return yaml.dump(
        payload,
        Dumper=_NoAliasDumper,
        sort_keys=False,
        allow_unicode=True,
        width=100,
        default_flow_style=False,
    )


def write_packages(
    package_ids: list[str],
    *,
    root: Path = REPO_ROOT,
    check: bool = False,
) -> list[Path]:
    """Write or drift-check the requested generated source packages."""
    changed: list[Path] = []
    for package_id in package_ids:
        relative_path, builder = PACKAGE_BUILDERS[package_id]
        package_path = root / relative_path
        rendered = render_package(builder(root))
        if package_path.exists() and package_path.read_text() == rendered:
            continue
        changed.append(relative_path)
        if not check:
            package_path.parent.mkdir(parents=True, exist_ok=True)
            package_path.write_text(rendered, encoding="utf-8")
    if check and changed:
        listing = ", ".join(str(path) for path in changed)
        raise RuntimeError(f"Generated source packages are stale: {listing}")
    return changed


def _load_yaml(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as file:
        payload = yaml.safe_load(file)
    if not isinstance(payload, dict):
        raise TypeError(f"Expected a YAML mapping in {path}")
    return payload


def _year_file(manifest: dict[str, Any], year: int) -> dict[str, Any]:
    files = manifest.get("files", {})
    entry = files.get(year, files.get(str(year)))
    if not isinstance(entry, dict):
        raise ValueError(f"Manifest has no file mapping for {year}")
    return entry


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    content = path.read_bytes()
    text = decode_delimited_text(content)
    reader = csv.DictReader(StringIO(text))
    headers = list(reader.fieldnames or ())
    if not headers:
        raise ValueError(f"CSV has no header: {path}")
    rows = [dict(row) for row in reader]
    return headers, rows


def _artifact_extracted_date(artifact: dict[str, Any]) -> str:
    fetched_at = artifact.get("fetched_at")
    if not isinstance(fetched_at, str) or len(fetched_at) < 10:
        raise ValueError("Artifact manifest must provide fetched_at")
    return fetched_at[:10]


def _require_headers(headers: list[str], *required: str) -> None:
    missing = [header for header in required if header not in headers]
    if missing:
        raise ValueError(f"Missing required source columns: {missing}")


def _require_numeric(row: dict[str, str], *columns: str) -> None:
    for column in columns:
        try:
            Decimal(row[column].strip())
        except (InvalidOperation, KeyError):
            raise ValueError(
                f"Expected numeric {column} for source row {row!r}"
            ) from None


def _county_fips(state_value: str, county_value: str) -> str:
    state_text = str(state_value).strip()
    county_text = str(county_value).strip()
    if not re.fullmatch(r"\d{1,2}", state_text):
        raise ValueError(f"State FIPS {state_value!r} is not a 1-2 digit code")
    if not re.fullmatch(r"\d{1,5}", county_text):
        raise ValueError(f"County FIPS {county_value!r} is not a 1-5 digit code")
    state_digits = state_text.zfill(2)
    county_digits = county_text.lstrip("0") or "0"
    if len(county_digits) > 3:
        full_fips = county_text.zfill(5)
        if not full_fips.startswith(state_digits):
            raise ValueError(
                f"County FIPS {county_value!r} does not match state {state_value!r}"
            )
        return full_fips
    return f"{state_digits}{county_digits.zfill(3)}"


def _selector_scalar(value: str) -> int | float | str:
    stripped = str(value).strip()
    if re.fullmatch(r"[+-]?\d+", stripped):
        return int(stripped)
    try:
        return float(stripped)
    except ValueError:
        return stripped


def _excel_column(number: int) -> str:
    name = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        name = f"{chr(65 + remainder)}{name}"
    return name


def _snap_month_period(value: Any) -> str:
    if not isinstance(value, str):
        raise ValueError(f"Expected SNAP month label, got {value!r}")
    try:
        parsed = datetime.strptime(value, "%b %Y")
    except ValueError as exc:
        raise ValueError(f"Unsupported SNAP month label: {value!r}") from exc
    return parsed.strftime("%Y-%m")


def main(argv: list[str] | None = None) -> int:
    """Generate or drift-check administrative source packages."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--package",
        action="append",
        choices=sorted(PACKAGE_BUILDERS),
        help="Package to generate; repeat for multiple packages (default: all).",
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Fail if checked-in generated packages differ from source artifacts.",
    )
    args = parser.parse_args(argv)
    package_ids = args.package or list(PACKAGE_BUILDERS)
    changed = write_packages(package_ids, check=args.check)
    if not args.check:
        for path in changed:
            print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
