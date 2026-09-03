#!/usr/bin/env python3
"""Build the Statbel fiscal-income-distribution extract and package metadata.

The eight publisher workbooks are immutable inputs. This script selects only
publisher cells, preserves workbook/sheet/cell coordinates and raw numeric
lexemes, and normalizes euro amounts to the requested cent representation.
It does not aggregate, reconcile, age, impute, or align publisher values.
"""

from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
import hashlib
import json
from pathlib import Path
import posixpath
import sys
from typing import Any
import xml.etree.ElementTree as ET
from zipfile import ZipFile

import openpyxl
from openpyxl.utils import get_column_letter
import yaml


PACKAGE_DIR = Path(__file__).resolve().parent
REPO_ROOT = PACKAGE_DIR.parents[2]
sys.path.insert(0, str(REPO_ROOT))

from chronicle.epoch import schema_id  # noqa: E402


PACKAGE_ID = "statbel-fiscal-income-distribution-2023"
SOURCE_PACKAGE_SCHEMA_VERSION = schema_id("source_package")
LANDING_PAGE = "https://statbel.fgov.be/en/themes/households/taxable-income"
RAW_BASE_URL = (
    "https://statbel.fgov.be/sites/default/files/files/documents/"
    "Huishoudens/10.9%20Fiscale%20inkomens"
)
EXTRACTED_AT = "2026-08-23"
INCOME_YEAR = 2023
ASSESSMENT_YEAR = 2024
CENT = Decimal("0.01")

DATA_DIR = REPO_ROOT / "db" / "data" / "statbel" / "fiscal_income_distribution_2023"
CSV_FILENAME = "statbel_fiscal_income_distribution_2023.csv"
CSV_PATH = DATA_DIR / CSV_FILENAME
MANIFEST_PATH = DATA_DIR / "manifest.yaml"
SOURCE_PACKAGE_PATH = PACKAGE_DIR / "source_package.yaml"

RAW_CODES = ("A_1", "A_2", "A_3", "B_1", "B_2", "B_3", "B_4", "B_5")


@dataclass(frozen=True)
class Geography:
    slug: str
    geography_id: str
    level: str
    name: str
    vintage: str
    sheet_name: str
    overview_row: int


GEOGRAPHIES = (
    Geography("be", "BE", "country", "Belgium", "current", "België", 8),
    Geography(
        "be2",
        "BE2",
        "nuts1",
        "Flemish Region",
        "NUTS_2024",
        "Vlaams Gewest",
        9,
    ),
    Geography(
        "be3",
        "BE3",
        "nuts1",
        "Walloon Region",
        "NUTS_2024",
        "Waals Gewest",
        10,
    ),
    Geography(
        "be1",
        "BE1",
        "nuts1",
        "Brussels Capital Region",
        "NUTS_2024",
        "Brussels Hoofdst. Gewest",
        11,
    ),
)

VALUE_METRICS = (
    "declarations",
    "declaration_share",
    "amount_eur",
    "amount_share",
    "upper_bound_eur",
    "taxable_income_eur",
    "taxable_income_share",
    "total_tax_eur",
    "total_tax_share",
    "average_tax_rate",
    "payable_tax_eur",
    "payable_tax_share",
    "tax_refund_eur",
    "tax_refund_share",
)

CSV_HEADERS = [
    "row_id",
    "table_code",
    "source_workbook",
    "dimension_1",
    "dimension_2",
    "dimension_3",
    "income_year",
    "assessment_year",
    "tax_return.net_taxable_income_class",
    "tax_return.net_taxable_income_lower_bound",
    "tax_return.net_taxable_income_upper_bound",
    "tax_return.rank_group",
    "tax_return.declaration_category",
    "tax_return.declaration_type",
    "tax_return.age_band",
    "tax_return.decile",
    "tax_return.dependants_group",
]
for geography in GEOGRAPHIES:
    CSV_HEADERS.extend(
        (
            f"{geography.slug}_source_sheet",
            f"{geography.slug}_source_row",
            f"{geography.slug}_source_cells",
            f"{geography.slug}_source_raw_values",
            *(f"{geography.slug}_{metric}" for metric in VALUE_METRICS),
        )
    )

CSV_COLUMNS = {
    header: get_column_letter(index)
    for index, header in enumerate(CSV_HEADERS, start=1)
}


def _raw_path(code: str) -> Path:
    return DATA_DIR / f"fisc2023_{code}_NL.xlsx"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _xlsx_numeric_lexemes(path: Path) -> dict[tuple[str, str], str]:
    """Read exact numeric ``<v>`` lexemes keyed by worksheet name and cell."""

    spreadsheet_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    document_rel_ns = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    with ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relationships = {
            relationship.attrib["Id"]: relationship.attrib["Target"]
            for relationship in relationships_root.findall(
                f"{{{package_rel_ns}}}Relationship"
            )
        }
        values: dict[tuple[str, str], str] = {}
        for sheet in workbook.findall(
            f"{{{spreadsheet_ns}}}sheets/{{{spreadsheet_ns}}}sheet"
        ):
            sheet_name = sheet.attrib["name"]
            relationship_id = sheet.attrib[f"{{{document_rel_ns}}}id"]
            target = relationships[relationship_id]
            if target.startswith("/"):
                archive_path = target.lstrip("/")
            else:
                archive_path = posixpath.normpath(posixpath.join("xl", target))
            worksheet = ET.fromstring(archive.read(archive_path))
            for cell in worksheet.iter(f"{{{spreadsheet_ns}}}c"):
                value = cell.find(f"{{{spreadsheet_ns}}}v")
                if value is None or value.text is None:
                    continue
                address = cell.attrib.get("r")
                if not address:
                    raise ValueError(
                        f"Cell without address in {path.name} {sheet_name}"
                    )
                values[(sheet_name, address)] = value.text
        return values


def _published_value(value: Any, kind: str) -> str | None:
    if _is_blank(value):
        return None
    if kind == "count":
        decimal = Decimal(str(value))
        if decimal != decimal.to_integral_value():
            raise ValueError(f"Expected integer count, got {value!r}")
        return str(int(decimal))
    decimal = Decimal(str(value))
    if kind == "eur":
        return format(decimal.quantize(CENT, rounding=ROUND_HALF_UP), "f")
    if kind == "ratio":
        return format(decimal, "f")
    raise ValueError(f"Unsupported publisher value kind: {kind}")


def _json_map(values: dict[str, str]) -> str:
    return json.dumps(values, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _slug_income_class(label: str) -> str:
    if label == "Totaal":
        return "total"
    if label == "Minder dan 1":
        return "class_under_1"
    if label == "100 en meer":
        return "class_100_plus"
    lower, upper = label.split("-", maxsplit=1)
    return f"class_{lower}_{upper}"


def _income_class_semantics(
    label: str, class_id: str
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    if class_id == "total":
        return {"tax_return.net_taxable_income_class": "all"}, []
    constraints: list[dict[str, Any]] = [
        {
            "variable": "tax_return.net_taxable_income_class",
            "operator": "==",
            "value": class_id,
            "label": f"Published Statbel €1,000 class {label}",
        }
    ]
    if class_id == "class_under_1":
        constraints.extend(
            (
                {
                    "variable": "tax_return.net_taxable_income",
                    "operator": ">",
                    "value": 0,
                    "unit": "eur",
                    "label": "Positive taxable income; zero-income returns are excluded",
                },
                {
                    "variable": "tax_return.net_taxable_income",
                    "operator": "<",
                    "value": 1000,
                    "unit": "eur",
                    "label": "Published class upper bound",
                },
            )
        )
    elif class_id == "class_100_plus":
        constraints.append(
            {
                "variable": "tax_return.net_taxable_income",
                "operator": ">=",
                "value": 100000,
                "unit": "eur",
                "label": "Published open-ended class lower bound",
            }
        )
    else:
        lower, upper = (int(part) * 1000 for part in label.split("-", maxsplit=1))
        constraints.extend(
            (
                {
                    "variable": "tax_return.net_taxable_income",
                    "operator": ">=",
                    "value": lower,
                    "unit": "eur",
                    "label": "Published class lower bound",
                },
                {
                    "variable": "tax_return.net_taxable_income",
                    "operator": "<",
                    "value": upper,
                    "unit": "eur",
                    "label": "Published class upper bound",
                },
            )
        )
    return {"tax_return.net_taxable_income_class": class_id}, constraints


def _eq_constraint(variable: str, value: Any, label: str) -> dict[str, Any]:
    return {"variable": variable, "operator": "==", "value": value, "label": label}


def _rank_id(decile: Any, percentile: Any = None) -> str:
    if decile == "Totaal":
        return "total"
    if not _is_blank(decile):
        return f"decile_{int(decile):02d}"
    if not _is_blank(percentile):
        return f"percentile_{int(percentile):03d}"
    raise ValueError(
        f"Missing rank label: decile={decile!r}, percentile={percentile!r}"
    )


def _rank_semantics(rank_id: str) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    value = "all" if rank_id == "total" else rank_id
    return (
        {"tax_return.rank_group": value},
        []
        if rank_id == "total"
        else [
            _eq_constraint(
                "tax_return.rank_group",
                rank_id,
                "Published Statbel decile or top-decile percentile",
            )
        ],
    )


def _age_semantics(age_id: str) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    if age_id == "total":
        return {"tax_return.age_band": "all"}, []
    return (
        {"tax_return.age_band": age_id},
        [
            _eq_constraint(
                "tax_return.age_band",
                age_id,
                "Exact published Statbel age category",
            )
        ],
    )


def _slug_age(label: str) -> str:
    if label == "Totaal":
        return "total"
    if label == "Minder dan 24 jaar":
        return "under_24"
    if label == "85 jaar en meer":
        return "age_85_plus"
    prefix = "Van "
    suffix = " jaar"
    if not label.startswith(prefix) or not label.endswith(suffix):
        raise ValueError(f"Unexpected B.4 age label: {label!r}")
    return "age_" + label.removeprefix(prefix).removesuffix(suffix).replace(
        " tot ", "_"
    )


def _new_csv_row(
    row_id: str,
    table_code: str,
    source_workbook: str,
    dimension_1: str = "",
    dimension_2: str = "",
    dimension_3: str = "",
) -> dict[str, str]:
    row = {header: "" for header in CSV_HEADERS}
    row.update(
        {
            "row_id": row_id,
            "table_code": table_code,
            "source_workbook": source_workbook,
            "dimension_1": dimension_1,
            "dimension_2": dimension_2,
            "dimension_3": dimension_3,
            "income_year": str(INCOME_YEAR),
            "assessment_year": str(ASSESSMENT_YEAR),
        }
    )
    return row


def _semantic_lexeme(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _apply_semantic_evidence(
    csv_row: dict[str, str],
    filters: dict[str, Any],
    constraints: list[dict[str, Any]],
) -> None:
    """Carry normalized publisher dimensions alongside the selected cells."""

    evidence: dict[str, Any] = dict(filters)
    for constraint in constraints:
        variable = str(constraint["variable"])
        operator = str(constraint["operator"])
        if operator == "==":
            header = variable
        elif operator in {">", ">="}:
            header = f"{variable}_lower_bound"
        elif operator in {"<", "<="}:
            header = f"{variable}_upper_bound"
        else:
            raise ValueError(f"Unsupported semantic-evidence operator: {operator}")
        value = constraint["value"]
        existing = evidence.get(header)
        if existing is not None and _semantic_lexeme(existing) != _semantic_lexeme(
            value
        ):
            raise ValueError(
                f"Conflicting semantic evidence for {csv_row['row_id']} {header}: "
                f"{existing!r} != {value!r}"
            )
        evidence[header] = value

    for header, value in evidence.items():
        if header not in csv_row:
            raise ValueError(
                f"Missing curated semantic column for {csv_row['row_id']}: {header}"
            )
        csv_row[header] = _semantic_lexeme(value)


def _fill_geography_values(
    csv_row: dict[str, str],
    geography: Geography,
    *,
    sheet_name: str,
    source_row: int,
    values: dict[str, tuple[Any, str, str]],
    source_lexemes: dict[tuple[str, str], str],
    seen_source_cells: set[tuple[str, str, str]],
) -> None:
    cells: dict[str, str] = {}
    raw_values: dict[str, str] = {}
    workbook = csv_row["source_workbook"]
    for metric, (raw_value, kind, address) in values.items():
        if _is_blank(raw_value):
            continue
        source_lexeme = source_lexemes.get((sheet_name, address))
        if source_lexeme is None:
            raise ValueError(
                f"Missing XLSX numeric lexeme: {workbook} {sheet_name}!{address}"
            )
        # Excel stores many already-rounded shares as binary-double artifacts
        # in XML. Preserve that XML lexeme in the provenance map, while using
        # openpyxl's decoded decimal spelling for the published ratio value.
        value_for_representation = raw_value if kind == "ratio" else source_lexeme
        published = _published_value(value_for_representation, kind)
        if published is None:
            raise ValueError(
                f"Blank XLSX numeric lexeme: {workbook} {sheet_name}!{address}"
            )
        csv_row[f"{geography.slug}_{metric}"] = published
        cells[metric] = address
        raw_values[metric] = source_lexeme
        source_key = (workbook, sheet_name, address)
        if source_key in seen_source_cells:
            raise ValueError(f"Publisher cell selected more than once: {source_key}")
        seen_source_cells.add(source_key)
    csv_row[f"{geography.slug}_source_sheet"] = sheet_name
    csv_row[f"{geography.slug}_source_row"] = str(source_row)
    csv_row[f"{geography.slug}_source_cells"] = _json_map(cells)
    csv_row[f"{geography.slug}_source_raw_values"] = _json_map(raw_values)


def _descriptor(
    *,
    row_id: str,
    value_id: str,
    label: str,
    filters: dict[str, Any] | None = None,
    constraints: list[dict[str, Any]] | None = None,
    table_record_kind: str = "detail",
) -> dict[str, Any]:
    return {
        "row_id": row_id,
        "value_id": value_id,
        "label": label,
        "filters": filters or {},
        "constraints": constraints or [],
        "table_record_kind": table_record_kind,
    }


def _extract_rows() -> tuple[
    list[dict[str, str]], dict[str, list[dict[str, Any]]], set[tuple[str, str, str]]
]:
    selected_codes = ("A_1", "B_1", "B_3", "B_4", "B_5")
    workbooks = {
        code: openpyxl.load_workbook(_raw_path(code), data_only=True, read_only=False)
        for code in selected_codes
    }
    source_lexemes = {
        code: _xlsx_numeric_lexemes(_raw_path(code)) for code in selected_codes
    }
    csv_rows: list[dict[str, str]] = []
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    seen_source_cells: set[tuple[str, str, str]] = set()

    # A.1 Home publishes zero-income declarations separately from the
    # positive-income distribution.
    workbook_name = _raw_path("A_1").name
    row = _new_csv_row(
        "a1.zero_income_declarations",
        "A_1_HOME",
        workbook_name,
        "zero_income_declarations",
    )
    home = workbooks["A_1"]["Home"]
    for geography in GEOGRAPHIES:
        _fill_geography_values(
            row,
            geography,
            sheet_name="Home",
            source_row=geography.overview_row,
            values={
                "declarations": (
                    home.cell(geography.overview_row, 2).value,
                    "count",
                    f"B{geography.overview_row}",
                )
            },
            source_lexemes=source_lexemes["A_1"],
            seen_source_cells=seen_source_cells,
        )
    csv_rows.append(row)
    groups["zero_income"].append(
        _descriptor(
            row_id=row["row_id"],
            value_id="total",
            label="Declarations with zero total net taxable income",
            table_record_kind="total",
        )
    )

    # A.1: six four-column components by 101 positive-income classes plus total.
    a1_components = (
        ("taxable_income", "Total net taxable income", 2),
        ("professional_income", "Total net professional income", 6),
        ("immovable_property_income", "Total net immovable-property income", 10),
        (
            "capital_and_movable_property_income",
            "Total net capital and movable-property income",
            14,
        ),
        ("miscellaneous_income", "Total net miscellaneous income", 18),
        ("deductible_expenditures", "Deductible expenditures", 22),
    )
    national_a1 = workbooks["A_1"][GEOGRAPHIES[0].sheet_name]
    for component_id, component_label, first_column in a1_components:
        group_id = f"a1.{component_id}"
        for source_row in range(9, 111):
            label = str(national_a1.cell(source_row, 1).value)
            class_id = _slug_income_class(label)
            row_id = f"a1.{component_id}.{class_id}"
            row = _new_csv_row(
                row_id,
                "A_1",
                workbook_name,
                component_id,
                class_id,
                label,
            )
            for geography in GEOGRAPHIES:
                sheet = workbooks["A_1"][geography.sheet_name]
                if str(sheet.cell(source_row, 1).value) != label:
                    raise ValueError(
                        f"A.1 row-label drift in {geography.sheet_name}!A{source_row}"
                    )
                metrics = (
                    ("declarations", first_column, "count"),
                    ("declaration_share", first_column + 1, "ratio"),
                    ("amount_eur", first_column + 2, "eur"),
                    ("amount_share", first_column + 3, "ratio"),
                )
                _fill_geography_values(
                    row,
                    geography,
                    sheet_name=geography.sheet_name,
                    source_row=source_row,
                    values={
                        metric: (
                            sheet.cell(source_row, column).value,
                            kind,
                            f"{get_column_letter(column)}{source_row}",
                        )
                        for metric, column, kind in metrics
                    },
                    source_lexemes=source_lexemes["A_1"],
                    seen_source_cells=seen_source_cells,
                )
            filters, constraints = _income_class_semantics(label, class_id)
            _apply_semantic_evidence(row, filters, constraints)
            csv_rows.append(row)
            groups[group_id].append(
                _descriptor(
                    row_id=row_id,
                    value_id=class_id,
                    label=f"{component_label}: {label}",
                    filters=filters,
                    constraints=constraints,
                    table_record_kind="total" if class_id == "total" else "detail",
                )
            )

    # B.1: published total, deciles 01-09, top-decile percentiles, and decile 10.
    workbook_name = _raw_path("B_1").name
    national_b1 = workbooks["B_1"][GEOGRAPHIES[0].sheet_name]
    b1_metrics = (
        ("upper_bound_eur", 3, "eur"),
        ("taxable_income_eur", 4, "eur"),
        ("taxable_income_share", 5, "ratio"),
        ("total_tax_eur", 6, "eur"),
        ("total_tax_share", 7, "ratio"),
        ("average_tax_rate", 8, "ratio"),
        ("payable_tax_eur", 9, "eur"),
        ("payable_tax_share", 10, "ratio"),
        ("tax_refund_eur", 11, "eur"),
        ("tax_refund_share", 12, "ratio"),
    )
    for source_row in range(8, 29):
        decile = national_b1.cell(source_row, 1).value
        percentile = national_b1.cell(source_row, 2).value
        rank_id = _rank_id(decile, percentile)
        publisher_label = decile if not _is_blank(decile) else percentile
        row_id = f"b1.{rank_id}"
        row = _new_csv_row(
            row_id,
            "B_1",
            workbook_name,
            rank_id,
            str(publisher_label),
        )
        for geography in GEOGRAPHIES:
            sheet = workbooks["B_1"][geography.sheet_name]
            other_rank = _rank_id(
                sheet.cell(source_row, 1).value,
                sheet.cell(source_row, 2).value,
            )
            if other_rank != rank_id:
                raise ValueError(
                    f"B.1 rank-label drift in {geography.sheet_name} row {source_row}"
                )
            _fill_geography_values(
                row,
                geography,
                sheet_name=geography.sheet_name,
                source_row=source_row,
                values={
                    metric: (
                        sheet.cell(source_row, column).value,
                        kind,
                        f"{get_column_letter(column)}{source_row}",
                    )
                    for metric, column, kind in b1_metrics
                },
                source_lexemes=source_lexemes["B_1"],
                seen_source_cells=seen_source_cells,
            )
        filters, constraints = _rank_semantics(rank_id)
        _apply_semantic_evidence(row, filters, constraints)
        csv_rows.append(row)
        descriptor = _descriptor(
            row_id=row_id,
            value_id=rank_id,
            label=f"B.1 {publisher_label}",
            filters=filters,
            constraints=constraints,
            table_record_kind="total" if rank_id == "total" else "detail",
        )
        partition = "percentile" if rank_id.startswith("percentile_") else "decile"
        groups[f"b1.{partition}_values"].append(descriptor)
        if row["be_upper_bound_eur"]:
            groups[f"b1.{partition}_bounds"].append(descriptor)

    # B.3: seven declaration categories by total/decile.
    workbook_name = _raw_path("B_3").name
    b3_categories = (
        (
            "individual_no_professional_income_man",
            "Individual, no professional income, man",
            2,
        ),
        (
            "individual_no_professional_income_woman",
            "Individual, no professional income, woman",
            4,
        ),
        (
            "individual_with_professional_income_man",
            "Individual, with professional income, man",
            6,
        ),
        (
            "individual_with_professional_income_woman",
            "Individual, with professional income, woman",
            8,
        ),
        ("joint_no_professional_income", "Joint, no professional income", 10),
        ("joint_one_professional_income", "Joint, one professional income", 12),
        ("joint_two_professional_incomes", "Joint, two professional incomes", 14),
    )
    national_b3 = workbooks["B_3"][GEOGRAPHIES[0].sheet_name]
    for category_id, category_label, first_column in b3_categories:
        for source_row in range(9, 20):
            raw_rank = national_b3.cell(source_row, 1).value
            rank_id = "total" if raw_rank == "Totaal" else f"decile_{int(raw_rank):02d}"
            row_id = f"b3.{category_id}.{rank_id}"
            row = _new_csv_row(
                row_id,
                "B_3",
                workbook_name,
                category_id,
                rank_id,
                str(raw_rank),
            )
            for geography in GEOGRAPHIES:
                sheet = workbooks["B_3"][geography.sheet_name]
                _fill_geography_values(
                    row,
                    geography,
                    sheet_name=geography.sheet_name,
                    source_row=source_row,
                    values={
                        "declarations": (
                            sheet.cell(source_row, first_column).value,
                            "count",
                            f"{get_column_letter(first_column)}{source_row}",
                        ),
                        "taxable_income_eur": (
                            sheet.cell(source_row, first_column + 1).value,
                            "eur",
                            f"{get_column_letter(first_column + 1)}{source_row}",
                        ),
                    },
                    source_lexemes=source_lexemes["B_3"],
                    seen_source_cells=seen_source_cells,
                )
            rank_filters, rank_constraints = _rank_semantics(rank_id)
            filters = {
                "tax_return.declaration_category": category_id,
                **rank_filters,
            }
            constraints = [
                _eq_constraint(
                    "tax_return.declaration_category",
                    category_id,
                    "Published B.3 declaration category",
                ),
                *rank_constraints,
            ]
            _apply_semantic_evidence(row, filters, constraints)
            csv_rows.append(row)
            groups["b3"].append(
                _descriptor(
                    row_id=row_id,
                    value_id=f"{category_id}.{rank_id}",
                    label=f"{category_label}, {rank_id}",
                    filters=filters,
                    constraints=constraints,
                    table_record_kind="total" if rank_id == "total" else "detail",
                )
            )

    # B.4: declaration type x exact publisher age category x decile.
    workbook_name = _raw_path("B_4").name
    b4_blocks = (
        ("individual", "Individual declaration", range(9, 24)),
        ("joint", "Joint declaration", range(25, 40)),
    )
    national_b4 = workbooks["B_4"][GEOGRAPHIES[0].sheet_name]
    for declaration_type, declaration_label, source_rows in b4_blocks:
        for source_row in source_rows:
            age_label = str(national_b4.cell(source_row, 2).value)
            age_id = _slug_age(age_label)
            for decile in range(1, 11):
                first_column = 3 + (decile - 1) * 2
                decile_id = f"decile_{decile:02d}"
                row_id = f"b4.{declaration_type}.{age_id}.{decile_id}"
                row = _new_csv_row(
                    row_id,
                    "B_4",
                    workbook_name,
                    declaration_type,
                    age_id,
                    decile_id,
                )
                for geography in GEOGRAPHIES:
                    sheet = workbooks["B_4"][geography.sheet_name]
                    if str(sheet.cell(source_row, 2).value) != age_label:
                        raise ValueError(
                            f"B.4 age-label drift in {geography.sheet_name} row {source_row}"
                        )
                    _fill_geography_values(
                        row,
                        geography,
                        sheet_name=geography.sheet_name,
                        source_row=source_row,
                        values={
                            "declarations": (
                                sheet.cell(source_row, first_column).value,
                                "count",
                                f"{get_column_letter(first_column)}{source_row}",
                            ),
                            "taxable_income_eur": (
                                sheet.cell(source_row, first_column + 1).value,
                                "eur",
                                f"{get_column_letter(first_column + 1)}{source_row}",
                            ),
                        },
                        source_lexemes=source_lexemes["B_4"],
                        seen_source_cells=seen_source_cells,
                    )
                age_filters, age_constraints = _age_semantics(age_id)
                filters = {
                    "tax_return.declaration_type": declaration_type,
                    **age_filters,
                    "tax_return.decile": decile,
                }
                constraints = [
                    _eq_constraint(
                        "tax_return.declaration_type",
                        declaration_type,
                        "Published declaration type",
                    ),
                    *age_constraints,
                    _eq_constraint(
                        "tax_return.decile",
                        decile,
                        "Published Statbel decile",
                    ),
                ]
                _apply_semantic_evidence(row, filters, constraints)
                csv_rows.append(row)
                groups["b4"].append(
                    _descriptor(
                        row_id=row_id,
                        value_id=f"{declaration_type}.{age_id}.{decile_id}",
                        label=(
                            f"{declaration_label}, {age_label}, decile {decile:02d}"
                        ),
                        filters=filters,
                        constraints=constraints,
                        table_record_kind="total" if age_id == "total" else "detail",
                    )
                )

    # B.5: declaration type x decile x published dependants category.
    workbook_name = _raw_path("B_5").name
    b5_blocks = (
        ("individual", "Individual declaration", range(9, 19)),
        ("joint", "Joint declaration", range(20, 30)),
    )
    dependant_groups = (
        ("dependants_0", "0 dependants", 0, 3),
        ("dependants_1", "1 dependant", 1, 5),
        ("dependants_2", "2 dependants", 2, 7),
        ("dependants_3", "3 dependants", 3, 9),
        ("dependants_4", "4 dependants", 4, 11),
        ("dependants_5_plus", "5 or more dependants", "5_plus", 13),
    )
    national_b5 = workbooks["B_5"][GEOGRAPHIES[0].sheet_name]
    for declaration_type, declaration_label, source_rows in b5_blocks:
        for source_row in source_rows:
            raw_decile = national_b5.cell(source_row, 2).value
            decile = int(raw_decile)
            decile_id = f"decile_{decile:02d}"
            for (
                dependant_id,
                dependant_label,
                dependant_value,
                first_column,
            ) in dependant_groups:
                row_id = f"b5.{declaration_type}.{decile_id}.{dependant_id}"
                row = _new_csv_row(
                    row_id,
                    "B_5",
                    workbook_name,
                    declaration_type,
                    decile_id,
                    dependant_id,
                )
                for geography in GEOGRAPHIES:
                    sheet = workbooks["B_5"][geography.sheet_name]
                    if int(sheet.cell(source_row, 2).value) != decile:
                        raise ValueError(
                            f"B.5 decile drift in {geography.sheet_name} row {source_row}"
                        )
                    _fill_geography_values(
                        row,
                        geography,
                        sheet_name=geography.sheet_name,
                        source_row=source_row,
                        values={
                            "declarations": (
                                sheet.cell(source_row, first_column).value,
                                "count",
                                f"{get_column_letter(first_column)}{source_row}",
                            ),
                            "taxable_income_eur": (
                                sheet.cell(source_row, first_column + 1).value,
                                "eur",
                                f"{get_column_letter(first_column + 1)}{source_row}",
                            ),
                        },
                        source_lexemes=source_lexemes["B_5"],
                        seen_source_cells=seen_source_cells,
                    )
                filters = {
                    "tax_return.declaration_type": declaration_type,
                    "tax_return.decile": decile,
                    "tax_return.dependants_group": dependant_value,
                }
                constraints = [
                    _eq_constraint(
                        "tax_return.declaration_type",
                        declaration_type,
                        "Published declaration type",
                    ),
                    _eq_constraint(
                        "tax_return.decile",
                        decile,
                        "Published Statbel decile",
                    ),
                    _eq_constraint(
                        "tax_return.dependants_group",
                        dependant_value,
                        "Published number-of-dependants category",
                    ),
                ]
                _apply_semantic_evidence(row, filters, constraints)
                csv_rows.append(row)
                groups["b5"].append(
                    _descriptor(
                        row_id=row_id,
                        value_id=f"{declaration_type}.{decile_id}.{dependant_id}",
                        label=(
                            f"{declaration_label}, decile {decile:02d}, "
                            f"{dependant_label}"
                        ),
                        filters=filters,
                        constraints=constraints,
                    )
                )

    _validate_extract(csv_rows, seen_source_cells)
    return csv_rows, dict(groups), seen_source_cells


def _as_decimal(value: str) -> Decimal:
    return Decimal(value)


def _validate_extract(
    rows: list[dict[str, str]], seen_source_cells: set[tuple[str, str, str]]
) -> None:
    by_id = {row["row_id"]: row for row in rows}
    if len(by_id) != len(rows):
        raise ValueError("Curated extract row IDs are not unique")

    expected_per_geography = {
        "A_1_HOME": 1,
        "A_1": 2448,
        "B_1": 207,
        "B_3": 154,
        "B_4": 600,
        "B_5": 240,
    }
    for geography in GEOGRAPHIES:
        actual: dict[str, int] = defaultdict(int)
        for row in rows:
            actual[row["table_code"]] += sum(
                bool(row[f"{geography.slug}_{metric}"]) for metric in VALUE_METRICS
            )
        if dict(actual) != expected_per_geography:
            raise ValueError(
                f"Unexpected fact-cell inventory for {geography.slug}: {dict(actual)}"
            )

    if len(seen_source_cells) != 3650 * len(GEOGRAPHIES):
        raise ValueError(
            f"Expected 14,600 unique fact cells, got {len(seen_source_cells)}"
        )

    for geography in GEOGRAPHIES:
        slug = geography.slug
        class_rows = [
            row
            for row in rows
            if row["table_code"] == "A_1"
            and row["dimension_1"] == "taxable_income"
            and row["dimension_2"] != "total"
        ]
        total = by_id["a1.taxable_income.total"]
        if sum(int(row[f"{slug}_declarations"]) for row in class_rows) != int(
            total[f"{slug}_declarations"]
        ):
            raise ValueError(f"A.1 class declarations do not sum for {slug}")
        if sum(
            _as_decimal(row[f"{slug}_amount_eur"]) for row in class_rows
        ) != _as_decimal(total[f"{slug}_amount_eur"]):
            raise ValueError(f"A.1 class amounts do not sum for {slug}")

        b1_total = by_id["b1.total"]
        b1_deciles = [by_id[f"b1.decile_{decile:02d}"] for decile in range(1, 11)]
        b1_percentiles = [
            by_id[f"b1.percentile_{percentile:03d}"] for percentile in range(91, 101)
        ]
        for metric in (
            "taxable_income_eur",
            "total_tax_eur",
            "payable_tax_eur",
            "tax_refund_eur",
        ):
            total_value = _as_decimal(b1_total[f"{slug}_{metric}"])
            if (
                sum(_as_decimal(row[f"{slug}_{metric}"]) for row in b1_deciles)
                != total_value
            ):
                raise ValueError(f"B.1 deciles do not sum for {slug} {metric}")
            if sum(
                _as_decimal(row[f"{slug}_{metric}"]) for row in b1_percentiles
            ) != _as_decimal(by_id["b1.decile_10"][f"{slug}_{metric}"]):
                raise ValueError(f"B.1 top percentiles do not sum for {slug} {metric}")

    zero_national = int(by_id["a1.zero_income_declarations"]["be_declarations"])
    zero_regions = sum(
        int(by_id["a1.zero_income_declarations"][f"{slug}_declarations"])
        for slug in ("be1", "be2", "be3")
    )
    if zero_regions != zero_national:
        raise ValueError("Regional zero-income declarations do not sum to Belgium")


def _write_csv(rows: list[dict[str, str]]) -> dict[str, int]:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", newline="", encoding="utf-8") as output:
        writer = csv.DictWriter(output, fieldnames=CSV_HEADERS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    return {row["row_id"]: line for line, row in enumerate(rows, start=2)}


def _yaml_rows(
    descriptors: list[dict[str, Any]], row_numbers: dict[str, int]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ordinal, descriptor in enumerate(descriptors):
        row = {
            "value_id": descriptor["value_id"],
            "label": descriptor["label"],
            "ordinal": ordinal,
            "row_number": row_numbers[descriptor["row_id"]],
            "expected_row_header_column": "A",
            "expected_row_header": descriptor["row_id"],
            "source_row_id": descriptor["row_id"],
            "table_record_kind": descriptor["table_record_kind"],
            "guard_cells": [
                {
                    "column": "B",
                    "expected_value": descriptor["row_id"]
                    .split(".", maxsplit=1)[0]
                    .upper()
                    .replace("A1", "A_1")
                    .replace("B1", "B_1")
                    .replace("B3", "B_3")
                    .replace("B4", "B_4")
                    .replace("B5", "B_5"),
                    "label": "publisher table",
                }
            ],
        }
        # The zero-income row uses the A_1 Home overview table label.
        if descriptor["row_id"] == "a1.zero_income_declarations":
            row["guard_cells"][0]["expected_value"] = "A_1_HOME"
        if descriptor["filters"]:
            row["filters"] = descriptor["filters"]
        if descriptor["constraints"]:
            row["constraints"] = descriptor["constraints"]
        rows.append(row)
    return rows


def _measure(
    geography: Geography,
    *,
    measure_id: str,
    label: str,
    ordinal: int,
    csv_metric: str,
    concept: str,
    source_concept: str,
    unit: str,
    aggregation: str,
    evidence_notes: str,
) -> dict[str, Any]:
    column_name = f"{geography.slug}_{csv_metric}"
    return {
        "measure_id": measure_id,
        "label": label,
        "ordinal": ordinal,
        "column": CSV_COLUMNS[column_name],
        "source_column_id": measure_id,
        "expected_column_header_row": 1,
        "expected_column_header": column_name,
        "concept": concept,
        "source_concept": source_concept,
        "concept_relation": "exact",
        "concept_authority": "ledger-be",
        "concept_evidence_url": LANDING_PAGE,
        "concept_evidence_notes": evidence_notes,
        "legal_vintage": "income_year_2023_assessment_year_2024",
        "unit": unit,
        "aggregation": aggregation,
        "expected_cell_type": "number",
    }


def _family(geography: Geography, suffix: str) -> str:
    root = "statbel.fiscal_income_distribution.income_year2023"
    if geography.slug == "be":
        return f"{root}.{suffix}"
    return f"{root}.regions.{geography.slug}.{suffix}"


def _record_set(
    geography: Geography,
    *,
    record_set_id: str,
    source_record_id_prefix: str,
    record_set_spec_id: str,
    groupby_dimension: str,
    rows: list[dict[str, Any]],
    measures: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "record_set_id": record_set_id,
        "provenance_class": "administrative",
        "record_set_spec_id": record_set_spec_id,
        "source_record_id_prefix": source_record_id_prefix,
        "sheet_name": "statbel_fiscal_income_distribution_2023",
        "period_type": "tax_year",
        "period": INCOME_YEAR,
        "geography_id": geography.geography_id,
        "geography_level": geography.level,
        "geography_name": geography.name,
        "geography_vintage": geography.vintage,
        "entity": "return",
        "entity_role": "personal_income_tax_return_unit",
        "domain": "personal_income_tax",
        "groupby_dimension": groupby_dimension,
        "shared_constraints": [
            _eq_constraint("income_year", INCOME_YEAR, "Publisher income year"),
            _eq_constraint(
                "assessment_year",
                ASSESSMENT_YEAR,
                "Publisher assessment year",
            ),
        ],
        "rows": rows,
        "measures": measures,
    }


def _a1_measures(geography: Geography, component_id: str) -> list[dict[str, Any]]:
    component_specs = {
        "taxable_income": (
            "Total net taxable income",
            "taxable_income_eur",
            "taxable_income_share",
            "belgium_pit_taxable_income",
            "statbel.fiscal_income.total_net_taxable_income",
        ),
        "professional_income": (
            "Total net professional income",
            "professional_income_eur",
            "professional_income_share",
            "statbel.fiscal_income.net_professional_income",
            "statbel.fiscal_income.total_net_professional_income",
        ),
        "immovable_property_income": (
            "Total net immovable-property income",
            "immovable_property_income_eur",
            "immovable_property_income_share",
            "statbel.fiscal_income.net_immovable_property_income",
            "statbel.fiscal_income.total_net_immovable_property_income",
        ),
        "capital_and_movable_property_income": (
            "Total net capital and movable-property income",
            "capital_and_movable_property_income_eur",
            "capital_and_movable_property_income_share",
            "statbel.fiscal_income.net_capital_and_movable_property_income",
            "statbel.fiscal_income.total_net_capital_and_movable_property_income",
        ),
        "miscellaneous_income": (
            "Total net miscellaneous income",
            "miscellaneous_income_eur",
            "miscellaneous_income_share",
            "statbel.fiscal_income.net_miscellaneous_income",
            "statbel.fiscal_income.total_net_miscellaneous_income",
        ),
        "deductible_expenditures": (
            "Deductible expenditures",
            "deductible_expenditures_eur",
            "deductible_expenditures_share",
            "statbel.fiscal_income.deductible_expenditures",
            "statbel.fiscal_income.deductible_expenditures",
        ),
    }
    label, amount_id, share_id, amount_concept, source_amount = component_specs[
        component_id
    ]
    notes = (
        f"Statbel A.1 publishes {label.lower()} by €1,000 class of total net "
        "taxable income. Returns with zero total net taxable income are excluded."
    )
    return [
        _measure(
            geography,
            measure_id="declarations",
            label=f"Declarations with {label.lower()}",
            ordinal=0,
            csv_metric="declarations",
            concept=f"statbel.fiscal_income.{component_id}_declaration_count",
            source_concept=f"statbel.fiscal_income.{component_id}.declaration_count",
            unit="count",
            aggregation="sum",
            evidence_notes=(
                notes
                + " A joint declaration is one tax-return unit; the published count is not a person count."
            ),
        ),
        _measure(
            geography,
            measure_id="declaration_share",
            label=f"Share of declarations with {label.lower()}",
            ordinal=1,
            csv_metric="declaration_share",
            concept=f"statbel.fiscal_income.{component_id}_declaration_share",
            source_concept=f"statbel.fiscal_income.{component_id}.declaration_share",
            unit="ratio",
            aggregation="share",
            evidence_notes=notes + " The workbook publishes the share as a fraction.",
        ),
        _measure(
            geography,
            measure_id=amount_id,
            label=label,
            ordinal=2,
            csv_metric="amount_eur",
            concept=amount_concept,
            source_concept=source_amount,
            unit="eur",
            aggregation="sum",
            evidence_notes=(
                notes
                + " Euro cells are represented to cents from the workbook's stored numeric value."
            ),
        ),
        _measure(
            geography,
            measure_id=share_id,
            label=f"Share of {label.lower()}",
            ordinal=3,
            csv_metric="amount_share",
            concept=f"statbel.fiscal_income.{component_id}_amount_share",
            source_concept=f"statbel.fiscal_income.{component_id}.amount_share",
            unit="ratio",
            aggregation="share",
            evidence_notes=notes + " The workbook publishes the share as a fraction.",
        ),
    ]


def _b1_measures(geography: Geography) -> list[dict[str, Any]]:
    universe_note = (
        "B.1 excludes declarations with zero total net taxable income; Statbel "
        "reports those declarations separately in A.1 Home. "
    )
    definitions = (
        (
            "taxable_income_eur",
            "Total net taxable income",
            "belgium_pit_taxable_income",
            "statbel.fiscal_income.total_net_taxable_income",
            "eur",
            "sum",
            "B.1 taxable-income mass for the published decile or percentile.",
        ),
        (
            "taxable_income_share",
            "Share of total net taxable income",
            "statbel.fiscal_income.taxable_income_share",
            "statbel.fiscal_income_distribution.b1.taxable_income_share",
            "ratio",
            "share",
            "B.1 publishes this share as a fraction.",
        ),
        (
            "total_tax_eur",
            "Total enrolled tax",
            "statbel.fiscal_income.enrolled_total_tax",
            "statbel.fiscal_income_distribution.b1.total_tax",
            "eur",
            "sum",
            "B.1 total enrolled tax; low-decile cells are negative because of refundable credits.",
        ),
        (
            "total_tax_share",
            "Share of total enrolled tax",
            "statbel.fiscal_income.enrolled_total_tax_share",
            "statbel.fiscal_income_distribution.b1.total_tax_share",
            "ratio",
            "share",
            "B.1 publishes this signed share as a fraction.",
        ),
        (
            "average_tax_rate",
            "Average assessment rate",
            "statbel.fiscal_income.average_assessment_rate",
            "statbel.fiscal_income_distribution.b1.average_assessment_rate",
            "ratio",
            "rate",
            "B.1 published average assessment rate, stored as a fraction.",
        ),
        (
            "payable_tax_eur",
            "Tax payable",
            "statbel.fiscal_income.tax_payable",
            "statbel.fiscal_income_distribution.b1.tax_payable",
            "eur",
            "sum",
            "B.1 published tax-to-pay amount.",
        ),
        (
            "payable_tax_share",
            "Share of tax payable",
            "statbel.fiscal_income.tax_payable_share",
            "statbel.fiscal_income_distribution.b1.tax_payable_share",
            "ratio",
            "share",
            "B.1 publishes this share as a fraction.",
        ),
        (
            "tax_refund_eur",
            "Tax refund",
            "statbel.fiscal_income.tax_refund",
            "statbel.fiscal_income_distribution.b1.tax_refund",
            "eur",
            "sum",
            "B.1 published tax-refund amount.",
        ),
        (
            "tax_refund_share",
            "Share of tax refunds",
            "statbel.fiscal_income.tax_refund_share",
            "statbel.fiscal_income_distribution.b1.tax_refund_share",
            "ratio",
            "share",
            "B.1 publishes this share as a fraction.",
        ),
    )
    return [
        _measure(
            geography,
            measure_id=measure_id,
            label=label,
            ordinal=ordinal,
            csv_metric=measure_id,
            concept=concept,
            source_concept=source_concept,
            unit=unit,
            aggregation=aggregation,
            evidence_notes=universe_note + notes,
        )
        for ordinal, (
            measure_id,
            label,
            concept,
            source_concept,
            unit,
            aggregation,
            notes,
        ) in enumerate(definitions)
    ]


def _count_amount_measures(
    geography: Geography, table_code: str, evidence_notes: str
) -> list[dict[str, Any]]:
    table_id = table_code.lower().replace("_", "")
    universe_note = (
        " The published table universe excludes declarations with zero total net "
        "taxable income; Statbel reports those declarations separately in A.1 Home."
    )
    return [
        _measure(
            geography,
            measure_id="declarations",
            label="Tax declarations",
            ordinal=0,
            csv_metric="declarations",
            concept="statbel.fiscal_income.declaration_count",
            source_concept=(
                f"statbel.fiscal_income_distribution.{table_id}.declaration_count"
            ),
            unit="count",
            aggregation="sum",
            evidence_notes=(
                evidence_notes
                + " A joint declaration is counted once; this is a tax-return-unit count."
                + universe_note
            ),
        ),
        _measure(
            geography,
            measure_id="taxable_income_eur",
            label="Total net taxable income",
            ordinal=1,
            csv_metric="taxable_income_eur",
            concept="belgium_pit_taxable_income",
            source_concept="statbel.fiscal_income.total_net_taxable_income",
            unit="eur",
            aggregation="sum",
            evidence_notes=(
                evidence_notes
                + " The concept matches the existing Statbel commune package; values are represented to cents."
                + universe_note
            ),
        ),
    ]


def _build_source_package(
    groups: dict[str, list[dict[str, Any]]], row_numbers: dict[str, int]
) -> int:
    yaml_rows = {
        group_id: _yaml_rows(descriptors, row_numbers)
        for group_id, descriptors in groups.items()
    }
    record_sets: list[dict[str, Any]] = []
    components = (
        "taxable_income",
        "professional_income",
        "immovable_property_income",
        "capital_and_movable_property_income",
        "miscellaneous_income",
        "deductible_expenditures",
    )
    for geography in GEOGRAPHIES:
        zero_family = _family(geography, "zero_income_declarations")
        record_sets.append(
            _record_set(
                geography,
                record_set_id=zero_family,
                source_record_id_prefix=zero_family,
                record_set_spec_id=(
                    "statbel.fiscal_income_distribution.zero_income_declarations.geography.v1"
                ),
                groupby_dimension="geography",
                rows=yaml_rows["zero_income"],
                measures=[
                    _measure(
                        geography,
                        measure_id="declarations",
                        label="Declarations with zero total net taxable income",
                        ordinal=0,
                        csv_metric="declarations",
                        concept="statbel.fiscal_income.zero_income_declaration_count",
                        source_concept=(
                            "statbel.fiscal_income.total_net_taxable_income.zero_declarations"
                        ),
                        unit="count",
                        aggregation="sum",
                        evidence_notes=(
                            "A.1 Home overview publishes declarations with zero total net taxable income separately; "
                            "the A.1 income-class table excludes them."
                        ),
                    )
                ],
            )
        )

        for component in components:
            family = _family(geography, f"{component}.by_income_class_eur1000")
            record_sets.append(
                _record_set(
                    geography,
                    record_set_id=family,
                    source_record_id_prefix=family,
                    record_set_spec_id=(
                        "statbel.fiscal_income_distribution.a1.by_income_class_eur1000.v1"
                    ),
                    groupby_dimension="net_taxable_income_class_eur1000",
                    rows=yaml_rows[f"a1.{component}"],
                    measures=_a1_measures(geography, component),
                )
            )

        b1_family = _family(geography, "by_decile")
        record_sets.append(
            _record_set(
                geography,
                record_set_id=f"{b1_family}.decile_values",
                source_record_id_prefix=b1_family,
                record_set_spec_id=(
                    "statbel.fiscal_income_distribution.b1.decile_values.v1"
                ),
                groupby_dimension="tax_return_decile_group",
                rows=yaml_rows["b1.decile_values"],
                measures=_b1_measures(geography),
            )
        )
        record_sets.append(
            _record_set(
                geography,
                record_set_id=f"{b1_family}.top_decile_percentile_values",
                source_record_id_prefix=b1_family,
                record_set_spec_id=(
                    "statbel.fiscal_income_distribution.b1.top_decile_percentile_values.v1"
                ),
                groupby_dimension="tax_return_top_decile_percentile_group",
                rows=yaml_rows["b1.percentile_values"],
                measures=_b1_measures(geography),
            )
        )
        record_sets.append(
            _record_set(
                geography,
                record_set_id=f"{b1_family}.decile_bounds",
                source_record_id_prefix=b1_family,
                record_set_spec_id=(
                    "statbel.fiscal_income_distribution.b1.decile_bounds.v1"
                ),
                groupby_dimension="tax_return_decile_group",
                rows=yaml_rows["b1.decile_bounds"],
                measures=[
                    _measure(
                        geography,
                        measure_id="upper_bound_eur",
                        label="Published income upper bound",
                        ordinal=0,
                        csv_metric="upper_bound_eur",
                        concept="statbel.fiscal_income.net_taxable_income_upper_bound",
                        source_concept=(
                            "statbel.fiscal_income_distribution.b1.upper_bound"
                        ),
                        unit="eur",
                        aggregation="quantile",
                        evidence_notes=(
                            "B.1 published decile bound. Declarations with zero total "
                            "net taxable income are excluded. Blank open-ended bounds "
                            "are omitted, not imputed."
                        ),
                    )
                ],
            )
        )
        record_sets.append(
            _record_set(
                geography,
                record_set_id=f"{b1_family}.top_decile_percentile_bounds",
                source_record_id_prefix=b1_family,
                record_set_spec_id=(
                    "statbel.fiscal_income_distribution.b1.top_decile_percentile_bounds.v1"
                ),
                groupby_dimension="tax_return_top_decile_percentile_group",
                rows=yaml_rows["b1.percentile_bounds"],
                measures=[
                    _measure(
                        geography,
                        measure_id="upper_bound_eur",
                        label="Published income upper bound",
                        ordinal=0,
                        csv_metric="upper_bound_eur",
                        concept="statbel.fiscal_income.net_taxable_income_upper_bound",
                        source_concept=(
                            "statbel.fiscal_income_distribution.b1.upper_bound"
                        ),
                        unit="eur",
                        aggregation="quantile",
                        evidence_notes=(
                            "B.1 published top-decile percentile bound. Declarations "
                            "with zero total net taxable income are excluded. Blank "
                            "open-ended bounds are omitted, not imputed."
                        ),
                    )
                ],
            )
        )

        b3_family = _family(geography, "declaration_type_professional_income.by_decile")
        record_sets.append(
            _record_set(
                geography,
                record_set_id=b3_family,
                source_record_id_prefix=b3_family,
                record_set_spec_id=(
                    "statbel.fiscal_income_distribution.b3.declaration_category_decile.v1"
                ),
                groupby_dimension="declaration_category_and_rank_group",
                rows=yaml_rows["b3"],
                measures=_count_amount_measures(
                    geography,
                    "B_3",
                    "B.3 publisher cell by declaration/professional-income category and decile.",
                ),
            )
        )

        b4_family = _family(geography, "declaration_type_age.by_decile")
        record_sets.append(
            _record_set(
                geography,
                record_set_id=b4_family,
                source_record_id_prefix=b4_family,
                record_set_spec_id=(
                    "statbel.fiscal_income_distribution.b4.declaration_type_age_decile.v1"
                ),
                groupby_dimension="declaration_type_age_and_decile",
                rows=yaml_rows["b4"],
                measures=_count_amount_measures(
                    geography,
                    "B_4",
                    "B.4 publisher cell by declaration type, exact published age category, and decile. "
                    "The first category is retained literally as 'Minder dan 24 jaar'; "
                    "its values match the publisher Home bands through age 24, so no "
                    "numeric age boundary is imposed.",
                ),
            )
        )

        b5_family = _family(geography, "declaration_type_dependants.by_decile")
        record_sets.append(
            _record_set(
                geography,
                record_set_id=b5_family,
                source_record_id_prefix=b5_family,
                record_set_spec_id=(
                    "statbel.fiscal_income_distribution.b5.declaration_type_dependants_decile.v1"
                ),
                groupby_dimension="declaration_type_dependants_and_decile",
                rows=yaml_rows["b5"],
                measures=_count_amount_measures(
                    geography,
                    "B_5",
                    "B.5 publisher cell by declaration type, decile, and published dependants category.",
                ),
            )
        )

    payload = {
        "schema_version": SOURCE_PACKAGE_SCHEMA_VERSION,
        "package_id": PACKAGE_ID,
        "label": (
            "Statbel fiscal income distribution 2023 by income class, decile, "
            "declaration type, age, and dependants"
        ),
        "artifact": {
            "source_name": "statbel_fiscal_income_distribution",
            "source_table": (
                "Statbel fiscal income distribution tables A.1, B.1, B.3, B.4, "
                "and B.5, income year 2023"
            ),
            "resource_package": "db",
            "resource_directory": "data/statbel/fiscal_income_distribution_2023",
            "manifest": "manifest.yaml",
            "vintage": "statbel_fiscal_income_distribution_income_year_2023",
            "extracted_at": EXTRACTED_AT,
            "extraction_method": (
                "deterministic openpyxl selection of publisher cells into a curated "
                "CSV; workbook, sheet, row, cell, and exact XLSX XML numeric lexemes "
                "are retained; euro representation is quantized directly from the XML "
                "numeric lexeme with Decimal and ROUND_HALF_UP; no source values are "
                "aggregated or reconciled"
            ),
            "parser": "delimited_text_full_rows",
            "artifact_year": INCOME_YEAR,
            "sheet_name": "statbel_fiscal_income_distribution_2023",
        },
        "record_sets": record_sets,
    }
    header = (
        "# Generated by packages/statbel/fiscal_income_distribution_2023/"
        "build_package.py.\n"
        "# Publisher workbooks and the curated extract are hash-pinned in the manifest.\n"
    )
    SOURCE_PACKAGE_PATH.write_text(
        header
        + yaml.safe_dump(
            payload,
            sort_keys=False,
            allow_unicode=True,
            width=100,
        ),
        encoding="utf-8",
    )
    return len(record_sets)


def _storage(filename: str, sha256: str) -> dict[str, Any]:
    key = f"raw/belgium/{PACKAGE_ID}/2023/{sha256}/{filename}"
    return {
        "r2": {
            "provider": "r2",
            "bucket": "ledger-raw",
            "key": key,
            "uri": f"r2://ledger-raw/{key}",
        }
    }


def _build_manifest() -> None:
    csv_sha = _sha256(CSV_PATH)
    files: dict[Any, Any] = {
        INCOME_YEAR: {
            "filename": CSV_FILENAME,
            "source_url": LANDING_PAGE,
            "source_table": (
                "Curated cell-level extract of Statbel A.1, B.1, B.3, B.4, and B.5"
            ),
            "sha256": csv_sha,
            "size_bytes": CSV_PATH.stat().st_size,
            "storage": _storage(CSV_FILENAME, csv_sha),
            "notes": (
                "Deterministic representation-only extract. Each row retains the "
                "publisher workbook, sheet, row, selected cell coordinates, and raw "
                "numeric lexemes. Monetary values are represented to cents; no values "
                "are aggregated, reconciled, aged, imputed, or aligned."
            ),
        }
    }
    for code in RAW_CODES:
        path = _raw_path(code)
        sha256 = _sha256(path)
        filename = path.name
        files[code] = {
            "filename": filename,
            "source_url": f"{RAW_BASE_URL}/{filename}",
            "source_table": f"Statbel fiscal income table {code}, income year 2023",
            "sha256": sha256,
            "size_bytes": path.stat().st_size,
            "storage": _storage(filename, sha256),
            "notes": (
                "Verbatim Statbel XLSX source capture. R2 location is declared only; "
                "upload occurs after lane review."
            ),
        }
    payload = {
        "source_id": "belgium",
        "package_id": PACKAGE_ID,
        "source_name": "Statbel fiscal income distribution 2023",
        "publisher": "Statbel",
        "source_page": LANDING_PAGE,
        "files": files,
    }
    MANIFEST_PATH.write_text(
        yaml.safe_dump(payload, sort_keys=False, allow_unicode=True, width=100),
        encoding="utf-8",
    )


def main() -> None:
    missing = [
        str(_raw_path(code)) for code in RAW_CODES if not _raw_path(code).exists()
    ]
    if missing:
        raise FileNotFoundError(f"Missing staged Statbel workbooks: {missing}")
    csv_rows, groups, seen_source_cells = _extract_rows()
    row_numbers = _write_csv(csv_rows)
    record_set_count = _build_source_package(groups, row_numbers)
    _build_manifest()
    print(
        json.dumps(
            {
                "csv_rows": len(csv_rows),
                "fact_cells": len(seen_source_cells),
                "record_sets": record_set_count,
                "csv_sha256": _sha256(CSV_PATH),
                "csv_size_bytes": CSV_PATH.stat().st_size,
            },
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
